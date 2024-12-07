from config.config import *
from gui.componentes import *
from gui.registro import *
from core.usuarios import *
from core.productos import *
from gui.registro import *
import csv
import json
import openpyxl  # Para exportar a excel
import os  #Para crear carpeta de exportaciones
from openpyxl.styles import Alignment, PatternFill, Font  # Estilos de excel
import tkinter.filedialog as fd


class Configuracion:
    def __init__(self, contenedor):
        self.contenedor = contenedor
        self.conexion = Database()
        self.frame_registro = None
        
        # -------------------------------- Configuración de frame --------------------------------
        self.frame_config = ctk.CTkScrollableFrame(master=self.contenedor, fg_color=COLOR_BG)
        self.frame_config.grid(sticky="nsew", padx=0)

        self.contenedor.grid_rowconfigure(0, weight=1)
        self.contenedor.grid_columnconfigure(0, weight=1)
        self.frame_config.grid_rowconfigure(2, weight=1)
        self.frame_config.grid_columnconfigure(0, weight=1)
        
        frame_contenido = ctk.CTkFrame(self.frame_config, fg_color=COLOR_BG)
        frame_contenido.grid(row=1, column=0, padx=120, pady=20, sticky="nsew")
        frame_contenido.grid_columnconfigure((0, 1), weight=1)  # Hacer que las columnas se expandan

        
        label_titulo = crear_label(frame_contenido, 
                                    metodo="grid", 
                                    text="Configuración", 
                                    font=("Roboto", 32, "bold"))
        label_titulo.grid(row=0, column=0, columnspan=2, pady=(20, 25), sticky="ew")
        
        
        # -------------------------------- Datos del usuario --------------------------------
        label_titulo = crear_label(frame_contenido, 
                                    metodo="grid", 
                                    text="Datos del usuario", 
                                    font=("Roboto", 24, "bold"))
        label_titulo.grid(row=1, column=0, columnspan=2, pady=0, sticky="ew")
        
        # Nombre
        label_nombre = crear_label(frame_contenido, 
                    text=f" Nombre y apellido", 
                    font=("Roboto", 18, "bold"),
                    metodo="grid", 
                    image=crear_imagen("src/assets/icons/user.png", size=(22, 22))
                    )
        label_nombre.grid(row=2, columnspan=2, pady=(10, 0), sticky="ew")
        
        nombre = crear_info(frame_contenido, 
                             text=f"{Usuario.usuario_actual[0][0].capitalize()} {Usuario.usuario_actual[0][1].capitalize()}",
                             metodo="grid",
                            )
        nombre.grid(row=3, columnspan=2, pady=0, sticky="ew")

        # Correo
        label_correo = crear_label(frame_contenido, 
                    text=f" Correo electrónico", 
                    font=("Roboto", 18, "bold"),
                    metodo="grid", 
                    image=crear_imagen("src/assets/icons/login-mail.png", size=(22, 22))
                    )
        label_correo.grid(row=4, column=0, pady=(15, 0), sticky="ew")
        
        correo = crear_info(frame_contenido, 
                             text=f"{Usuario.usuario_actual[0][4]}",
                             metodo="grid",
                            )
        correo.grid(row=5, column=0, pady=0, padx=(0, 10), sticky="ew")

        # DNI
        label_dni = crear_label(frame_contenido, 
                    text=f" D.N.I", 
                    font=("Roboto", 18, "bold"),
                    metodo="grid",
                    image=crear_imagen("src/assets/icons/id.png", size=(22, 22))
                    )
        label_dni.grid(row=4, column=1, pady=(15, 0), sticky="ew")
        
        dni = crear_info(frame_contenido, 
                             text=f"{Usuario.usuario_actual[0][5]}",
                             metodo="grid",
                            )
        dni.grid(row=5, column=1, pady=0, padx=(10, 0), sticky="ew")
        
        
        # -------------------------------- Restablecer contraseña --------------------------------
        label_correo = crear_label(frame_contenido, 
                                   text=f"Restablecer contraseña",
                                   font=("Roboto", 24, "bold"),
                                   metodo="grid",
                                   )
        label_correo.grid(row=6, columnspan=2, pady=(35, 0), sticky="ew")
        
        
        label_correo = crear_label(frame_contenido, 
                                   text=f" Contraseña actual",
                                   font=("Roboto", 18, "bold"),
                                   metodo="grid",
                                   image=crear_imagen("src/assets/icons/login-password.png", size=(22, 22))
                                   )
        label_correo.grid(row=7, column=0, pady=(10, 0), sticky="ew")
        
        entry_contrasena_actual = crear_entry(frame_contenido, 
                            placeholder_text="**********", 
                            metodo="grid")
        entry_contrasena_actual.grid(row=8, column=0, pady=0, sticky="ew", padx=(0, 10))
        
        
        label_contrasena_nueva = crear_label(frame_contenido, 
                                   text=f" Nueva contraseña",
                                   font=("Roboto", 18, "bold"),
                                   metodo="grid",
                                   image=crear_imagen("src/assets/icons/login-password.png", size=(22, 22))
                                   )
        label_contrasena_nueva.grid(row=7, column=1, pady=(10, 0), sticky="ew")
        
        entry_contrasena_nueva = crear_entry(frame_contenido, 
                            placeholder_text="**********", 
                            metodo="grid")
        entry_contrasena_nueva.grid(row=8, column=1, pady=0, sticky="ew", padx=(10, 0))
        
        
        btn_contrasena = crear_boton(frame_contenido,
                                     metodo="grid", 
                                     text="Actualizar contraseña")
        btn_contrasena.grid(row=9, columnspan=2, pady=(20, 0), sticky="ew")  
            
        # -------------------------------- Exportar productos --------------------------------
        label_exportar = crear_label(frame_contenido, 
                                     text="Exportar productos", 
                                     font=("Roboto", 24, "bold"), 
                                     metodo="grid")
        label_exportar.grid(row=10, columnspan=2, pady=(35, 5), sticky="ew")
        
        exportar_optionmenu = crear_optionmenu(
            parent=frame_contenido,
            values=["Seleccione formato", "Excel", "CSV", "JSON"],
            pady=0,
            metodo="grid",
            )
        exportar_optionmenu.grid(row=11, column=0, sticky="ew", padx=(0, 10))
        
        exportar_btn = crear_boton(parent=frame_contenido, 
                                   text="Exportar", 
                                   metodo="grid",
                                   command=lambda: self.exportar_productos(exportar_optionmenu.get())
                                   )
        exportar_btn.grid(row=11, column=1, sticky="ew", padx=(10, 0))
        
        # -------------------------------- Importar productos --------------------------------
        label_importar = crear_label(frame_contenido, 
                                     text="Importar productos", 
                                     font=("Roboto", 24, "bold"), 
                                     metodo="grid")
        label_importar.grid(row=12, columnspan=2, pady=(35, 5), sticky="ew")
        
        importar_optionmenu = crear_optionmenu(
            parent=frame_contenido,
            values=["Seleccione formato", "Excel", "CSV", "JSON"],
            pady=0,
            metodo="grid",
            )
        importar_optionmenu.grid(row=13, column=0, sticky="ew", padx=(0, 10), pady=0)
        
        importar_btn = crear_boton(parent=frame_contenido, 
                                   text="Importar", 
                                   metodo="grid",
                                   command=lambda: self.importar_productos(importar_optionmenu.get())
                                   )
        importar_btn.grid(row=13, column=1, sticky="ew", padx=(10, 0), pady=0)
    
        
        # -------------------------------- MENÚ SUPERVISOR --------------------------------
        if Usuario.usuario_actual[0][2] == "supervisor":
            label_agregar_empleado = crear_label(frame_contenido, 
                                        text="Gestionar empleados", 
                                        font=("Roboto", 24, "bold"), 
                                        metodo="grid")
            label_agregar_empleado.grid(row=14, columnspan=2, pady=(35, 0), sticky="ew")
            
            # -------------------------------- Registrar empleado --------------------------------
            label_registrar_empleado = crear_label(frame_contenido, 
                                    text=f" Registrar empleado",
                                    font=("Roboto", 18, "bold"),
                                    metodo="grid",
                                    image=crear_imagen("src/assets/icons/user.png", size=(22, 22))
                                    )
            label_registrar_empleado.grid(row=15, column=0, pady=(10, 0), sticky="ew")
            
            btn_registrar_empleado= crear_boton(frame_contenido,
                                        metodo="grid", 
                                        text="Registrar empleado",
                                        command=self.mostrar_registro)
            btn_registrar_empleado.grid(row=16, column=0, pady=(0, 10), padx=(0, 10), sticky="ew")
            
            # -------------------------------- Otorgar rol --------------------------------
            label_otorgar_rol = crear_label(frame_contenido, 
                                    text=f" Otorgar rol",
                                    font=("Roboto", 18, "bold"),
                                    metodo="grid",
                                    image=crear_imagen("src/assets/icons/pencil.png", size=(22, 22))
                                    )
            label_otorgar_rol.grid(row=15, column=1, pady=(10, 0), padx=(0, 10), sticky="ew")
            
            btn_otorgar_rol= crear_boton(frame_contenido,
                                        metodo="grid", 
                                        text="Cambiar rol",
                                        command=self.mostrar_otorgar_rol)
            btn_otorgar_rol.grid(row=16, column=1, pady=(0, 10), padx=(10, 0), sticky="ew")
            
    # -------------------------------- Exportar --------------------------------
    def exportar_productos(self, formato):
        if formato == "CSV":
            self.exportar_csv()
        elif formato == "JSON":
            self.exportar_json()
        elif formato == "Excel":
            self.exportar_excel()
        else:
            crear_notificacion(self.contenedor, "error", "Seleccione un formato.")
    
    def exportar_csv(self):
        productos = self.conexion.ejecutar_bd("SELECT * FROM productos", tipo="select")
        lotes = self.conexion.ejecutar_bd("SELECT * FROM lotes", tipo="select")
        
        os.makedirs("Exportaciones", exist_ok=True)  # Crea carpeta para exportaciones
        
        with open('Exportaciones/productos.csv', mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["ID", "Nombre", "Marca", "Categoría", "Precio de compra", "Precio de venta"])
            for producto in productos:
                writer.writerow(producto)
        
        with open('Exportaciones/lotes.csv', mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Lote", "Producto ID", "Cantidad", "Fecha de vencimiento"])
            for lote in lotes:
                writer.writerow(lote)
        
        crear_notificacion(self.contenedor, "info", "Productos exportados a CSV correctamente.")
    
    def exportar_json(self):
        productos_lista = []
        productos = self.conexion.ejecutar_bd("SELECT * FROM productos", tipo="select")
        
        for producto in productos:
            productos_lista.append({
                "ID": producto[0],
                "Nombre": producto[1],
                "Marca": producto[2],
                "Categoría": producto[3],
                "Precio de compra": float(producto[4]),
                "Precio de venta": float(producto[5])
            })
            
        lotes_lista = []
        lotes = self.conexion.ejecutar_bd("SELECT * FROM lotes", tipo="select")
        
        for lote in lotes:
            lotes_lista.append({
                "Lote": lote[0],
                "Producto ID": lote[1],
                "Cantidad": lote[2],
                "Fecha de vencimiento": str(lote[3]),
            })
        
        os.makedirs("Exportaciones", exist_ok=True)  # Crea carpeta para exportaciones
        
        # Guardar en archivo
        with open('Exportaciones/productos.json', 'w', encoding='utf-8') as file:
            json.dump(productos_lista, file, ensure_ascii=False, indent=4)
        
        with open('Exportaciones/lotes.json', 'w', encoding='utf-8') as file:
            json.dump(lotes_lista, file, ensure_ascii=False, indent=4)
        
        crear_notificacion(self.contenedor, "info", "Productos exportados a JSON correctamente.")
    
    def exportar_excel(self):
        try:
            productos = self.conexion.ejecutar_bd("SELECT * FROM productos", tipo="select")
            lotes = self.conexion.ejecutar_bd("SELECT * FROM lotes", tipo="select")

            wb = openpyxl.Workbook()  # Contenedor para generar excel

            # ----------------- Hoja de Productos -----------------
            ws_productos = wb.active
            ws_productos.title = "Productos"

            # Encabezados
            headers_productos = ["ID", "Nombre", "Marca", "Categoría", "Precio de compra", "Precio de venta"]
            ws_productos.append(headers_productos)

            # Agregar datos de productos
            for producto in productos:
                ws_productos.append(producto)

            for col, header in enumerate(headers_productos, start=1):
                # Ajustar ancho de columnas
                max_length = max((len(str(row[col - 1])) for row in productos), default=len(header))
                ancho_ajustado = max(len(header), max_length) + 2
                ws_productos.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho_ajustado
                
                # Estilo de los encabezados
                celda = ws_productos.cell(row=1, column=col)
                celda.fill = PatternFill(start_color="0c955a", end_color="0c955a", fill_type="solid")
                celda.font = Font(bold=True, color="FFFFFF")
                celda.alignment = Alignment(horizontal="center", vertical="center")
            
            
            # ----------------- Hoja de Lotes -----------------
            ws_lotes = wb.create_sheet(title="Lotes")

            # Encabezados
            headers_lotes = ["Lote", "Producto ID", "Cantidad", "Fecha de vencimiento"]
            ws_lotes.append(headers_lotes)

            # Agregar datos de lotes
            for lote in lotes:
                ws_lotes.append(lote)
                
            for col, header in enumerate(headers_lotes, start=1):
                # Ajustar ancho de columnas
                max_length = max((len(str(row[col - 1])) for row in lotes), default=len(header))
                ancho_ajustado = max(len(header), max_length) + 2
                ws_lotes.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho_ajustado
                
                # Estilo de los encabezados
                celda = ws_lotes.cell(row=1, column=col)
                celda.fill = PatternFill(start_color="0c955a", end_color="0c955a", fill_type="solid")
                celda.font = Font(bold=True, color="FFFFFF")
                celda.alignment = Alignment(horizontal="center", vertical="center")
            
            os.makedirs("Exportaciones", exist_ok=True)  # Crea carpeta para exportaciones
            
            # Guardar archivo
            excel_file = "Exportaciones/productos_y_lotes.xlsx"
            wb.save(excel_file)
            
            crear_notificacion(self.contenedor, "info", "Productos exportados a Excel correctamente.")
        
        except Exception as e:
            print(f"Error al exportar a Excel: {e}")
            crear_notificacion(self.contenedor, "error", "El archivo ya existe.")
        
            
    # -------------------------------- Importar --------------------------------
    def importar_productos(self, formato):
        if formato == "CSV":
            self.importar_csv()
        elif formato == "JSON":
            self.importar_json()
        elif formato == "Excel":
            self.importar_excel()
        else:
            crear_notificacion(self.contenedor, "error", "Seleccione un formato.")
    
    def importar_json(self):
        # Ventana para seleccionar archivo de productos
        archivo_productos = fd.askopenfilename(title="Seleccionar archivo productos.json", 
                                              filetypes=[("Archivos JSON", "*.json")])
        
        # Ventana para seleccionar archivo de lotes
        archivo_lotes = fd.askopenfilename(title="Seleccionar archivo lotes.json", 
                                           filetypes=[("Archivos JSON", "*.json")])

        if archivo_productos and archivo_lotes:
            try:
                # Cargar productos desde archivo JSON
                with open(archivo_productos, 'r', encoding='utf-8') as file:
                    productos_data = json.load(file)
                    if isinstance(productos_data, list):
                        productos = productos_data  # Es una lista de productos directamente
                        for producto in productos:
                            query = """
                                INSERT INTO productos (id, nombre, marca, categoria, precio_compra, precio_venta) 
                                VALUES (%s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                    nombre = VALUES(nombre),
                                    marca = VALUES(marca),
                                    categoria = VALUES(categoria),
                                    precio_compra = VALUES(precio_compra),
                                    precio_venta = VALUES(precio_venta)
                            """
                            self.conexion.ejecutar_bd(query, tipo="insert", valores=(
                                producto['ID'],
                                producto['Nombre'],
                                producto['Marca'],
                                producto['Categoría'],
                                producto['Precio de compra'],
                                producto['Precio de venta']
                            ))
                    else:
                        crear_notificacion(self.contenedor, "error", "El archivo de productos no tiene el formato esperado (debe ser una lista de objetos).")

                # Cargar lotes desde archivo JSON
                with open(archivo_lotes, 'r', encoding='utf-8') as file:
                    lotes_data = json.load(file)
                    if isinstance(lotes_data, list):
                        lotes = lotes_data  # Es una lista de lotes directamente
                        for lote in lotes:
                            query = """
                                INSERT INTO lotes (lote, producto_id, cantidad, fecha_vencimiento) 
                                VALUES (%s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                    lote = VALUES(lote),
                                    producto_id = VALUES(producto_id),
                                    cantidad = VALUES(cantidad),
                                    fecha_vencimiento = VALUES(fecha_vencimiento)
                            """
                            self.conexion.ejecutar_bd(query, tipo="insert", valores=(
                                lote['Lote'],
                                lote['Producto ID'],
                                lote['Cantidad'],
                                lote['Fecha de vencimiento']
                            ))
                    else:
                        crear_notificacion(self.contenedor, "error", "El archivo de lotes no tiene el formato esperado (debe ser una lista de objetos).")

                crear_notificacion(self.contenedor, "info", "Productos y lotes importados correctamente.")
            except Exception as e:
                crear_notificacion(self.contenedor, "error", f"Error: Seleccione los archivos correctos.")
        else:
            crear_notificacion(self.contenedor, "error", "Debe seleccionar los archivos 'productos' y 'lotes'.")
    
    
    # -------------------------------- Agregar empleado --------------------------------
    def mostrar_registro(self):
        # Muestra el frame de registro y oculta el frame principal
        self.frame_registro = RegistroFrame(
            master=self.contenedor,
            frame_cambiar=self.regresar_configuracion
        )

        self.frame_config.grid_forget()

        self.frame_registro.grid(sticky="nsew")

    def regresar_configuracion(self):
        if self.frame_registro:
            self.frame_registro.grid_forget()
        # Mostrar nuevamente el frame de configuración
        self.frame_config.grid(sticky="nsew")
    
    # -------------------------------- Otorgar rol --------------------------------
    def mostrar_otorgar_rol(self):
        self.frame_config.grid_forget()
        
        self.frame_otorgar_rol = ctk.CTkFrame(master=self.contenedor, fg_color=COLOR_BG)
        self.frame_otorgar_rol.pack(expand=True, fill="x", padx=130)
     
        datos = self.conexion.ejecutar_bd(
            """
            SELECT documento, nombre, apellido 
            FROM usuarios
            """
        )
        
        empleados = [f"{documento} - {nombre} {apellido}" for documento, nombre, apellido in datos]

        crear_label(self.frame_otorgar_rol,
                    text="Otorgar rol", 
                    font=("Roboto", 32, "bold")
        )

        crear_label(self.frame_otorgar_rol,
                    text=" Seleccionar usuario",
                    font=("Roboto", 18, "bold"),
                    image=crear_imagen("src/assets/icons/user.png", size=(22, 22))
        )
        
        self.usuario_empleados = crear_optionmenu(
                    parent=self.frame_otorgar_rol,
                    values=["Seleccionar empleado"] + empleados,
                    pady=0
        )
        
        crear_label(self.frame_otorgar_rol,
                    text="Ó",
                    font=("Roboto", 18, "bold"),
                    pady= (15, 0),
                    anchor="center",
        )
        
        crear_label(self.frame_otorgar_rol,
                    text=" Asignar por D.N.I",
                    font=("Roboto", 18, "bold"),
                    image=crear_imagen("src/assets/icons/id.png", size=(22, 22))
        )
        
        self.documento_input = crear_entry(self.frame_otorgar_rol, fill="x", pady=0)

        crear_label(self.frame_otorgar_rol,
                    text=" Rol", 
                    font=("Roboto", 18, "bold"),
                    image=crear_imagen("src/assets/icons/pencil.png", size=(22, 22)),
                    pady=(20, 10)
        )

        self.usuario_rol = crear_optionmenu(
            parent=self.frame_otorgar_rol,
            values=["Empleado", "Supervisor"],
            pady=0
        )

        botones_frame = ctk.CTkFrame(self.frame_otorgar_rol, fg_color=COLOR_BG)
        botones_frame.pack(pady=(20, 0), fill="x")

        btn_guardar = crear_boton(botones_frame, text="Guardar", command=lambda: self.guardar_rol(self.contenedor), metodo="grid")
        btn_guardar.pack(side="left", padx=(0, 10), fill="x", expand="True")

        btn_cancelar = crear_boton(botones_frame, text="Cancelar", command=self.cancelar_rol, metodo="grid")
        btn_cancelar.pack(side="right", padx=(10, 0), fill="x", expand="True")
        
    def cancelar_rol(self):
        self.frame_otorgar_rol.pack_forget()
        self.frame_config.grid(sticky="nsew")

    def guardar_rol(self, parent_frame):
        # Validar si se ingresó un DNI por escrito
        documento = self.documento_input.get().strip()
        
        if not documento:  # Si el campo está vacío
            # Validar que se seleccionó un empleado en el optionmenu
            empleado_seleccionado = self.usuario_empleados.get()
            if empleado_seleccionado == "Seleccionar empleado":
                crear_notificacion(parent_frame, "error", "Selecciona un empleado o ingresa un DNI.")
                return
            
            # Extraer el documento del empleado seleccionado (antes del primer guion)
            documento = empleado_seleccionado.split(" - ")[0]

        if not documento.isdigit():
            crear_notificacion(parent_frame, "error", "El DNI debe contener solo números.")
            return
        
        # Validar si el documento existe en la base de datos
        existe_documento = self.conexion.ejecutar_bd(
            """
            SELECT COUNT(*) 
            FROM usuarios 
            WHERE documento = %s
            """,
            (documento,)
        )[0][0]

        if not existe_documento:
            crear_notificacion(parent_frame, "error", f"El documento no es válido.")
            return

        rol_seleccionado = self.usuario_rol.get()
        
        try:
            self.conexion.ejecutar_bd(
                """
                UPDATE usuarios
                SET rol = %s
                WHERE documento = %s
                """, 
                (rol_seleccionado.lower(), documento), 
                "update"
            )
            
            crear_notificacion(parent_frame, "info", f"Se asignó el rol '{rol_seleccionado}' al usuario.")
            
            # Limpiar los campos después de guardar
            self.usuario_empleados.set("Seleccionar empleado")
            self.usuario_rol.set("Empleado")
            self.documento_input.delete(0, "end")
            
        except Exception as e:
            print(f"No se pudo asignar el rol: {e}")
    